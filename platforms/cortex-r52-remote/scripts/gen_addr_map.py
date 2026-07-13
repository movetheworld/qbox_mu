#!/usr/bin/env python3
# Copyright (c) 2025 Qualcomm Innovation Center, Inc. All Rights Reserved.
# SPDX-License-Identifier: BSD-3-Clause
"""Generate addr-map.xlsx from the platform memory map."""

from pathlib import Path

ROWS = [
    ["Name", "Start Address", "End Address", "Size", "Type", "Access", "GIC INTID", "Description"],
    ["RAM / Firmware", "0x00000000", "0x0001FFFF", "128 KiB", "Memory", "RWX", "-", "Boot ROM/RAM, firmware image loaded here"],
    ["GIC Distributor (GICD)", "0x2F000000", "0x2F00FFFF", "64 KiB", "GICv3", "RW", "-", "GIC distributor registers"],
    ["GIC Redistributor (GICR)", "0x2F100000", "0x2F11FFFF", "128 KiB", "GICv3", "RW", "-", "CPU0 redistributor (SGI/PPI frame at +0x10000)"],
    ["PL011 UART", "0xC0000000", "0xC0000FFF", "4 KiB", "UART", "RW", "32 (SPI 0)", "Console UART, TX/RX via char_backend_stdio"],
    ["IRQ Test Generator", "0xC0001000", "0xC0001FFF", "4 KiB", "MMIO", "RW", "33 (SPI 1)", "Write 0x4=1 to start pulse; write 0x0=1 to clear"],
    ["Arch Timer PPI", "-", "-", "-", "GIC PPI", "-", "30", "Physical non-secure EL1 timer (wired in VP, optional in firmware)"],
]

def main():
    out_dir = Path(__file__).resolve().parent.parent
    xlsx_path = out_dir / "addr-map.xlsx"

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        csv_path = out_dir / "addr-map.csv"
        with csv_path.open("w", encoding="utf-8", newline="") as f:
            for row in ROWS:
                f.write(",".join(row) + "\n")
        print(f"openpyxl not installed; wrote {csv_path}")
        print("Install with: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Cortex-R52 Address Map"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for r, row in enumerate(ROWS, start=1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            if r == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 48)

    wb.save(xlsx_path)
    print(f"Wrote {xlsx_path}")


if __name__ == "__main__":
    main()
